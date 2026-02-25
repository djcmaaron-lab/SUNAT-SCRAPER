from flask import Flask, jsonify, request, send_file
import threading, time, csv, io, json
from enricher import enriquecer_lote, get_rubros_ccl

app = Flask(__name__)

STATE = {
    "running": False, "finished": False, "error": None,
    "progress": 0, "index": 0, "total": 0,
    "empresa_actual": "", "con_rep": 0, "con_tel": 0, "con_email": 0,
    "logs": [], "results": [], "started_at": None, "finished_at": None,
}
_LOCK = threading.RLock()

HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Lead Enricher · Perú</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*{margin:0;padding:0;box-sizing:border-box}

:root{
  --bg:#f0efe8;
  --surface:#ffffff;
  --surface2:#f7f6f0;
  --border:#e0ddd4;
  --border2:#ccc9be;
  --text:#1a1916;
  --muted:#7a7769;
  --accent:#d4500a;
  --accent-bg:#fdf1eb;
  --green:#1a7a4a;
  --green-bg:#edf7f2;
  --blue:#1a5fa0;
  --blue-bg:#edf3fb;
  --font:'DM Sans',sans-serif;
  --mono:'DM Mono',monospace;
}

body{background:var(--bg);color:var(--text);font-family:var(--font);min-height:100vh}

/* ── LAYOUT ── */
.shell{display:flex;min-height:100vh}

/* sidebar */
.sidebar{
  width:260px;flex-shrink:0;
  background:var(--text);color:#f0efe8;
  padding:2rem 1.5rem;
  display:flex;flex-direction:column;
  position:sticky;top:0;height:100vh;overflow:hidden;
}
.logo{font-size:1.15rem;font-weight:600;letter-spacing:-0.02em;margin-bottom:0.3rem}
.logo span{color:#d4500a}
.logo-sub{font-size:0.72rem;color:#7a7769;font-family:var(--mono);letter-spacing:0.05em;margin-bottom:2.5rem}

.steps{flex:1}
.step-item{
  display:flex;align-items:flex-start;gap:0.75rem;
  padding:0.6rem 0.75rem;border-radius:8px;
  cursor:pointer;transition:background 0.15s;
  margin-bottom:0.2rem;
}
.step-item:hover{background:rgba(255,255,255,0.06)}
.step-item.active{background:rgba(212,80,10,0.2)}
.step-item.done .step-num{background:var(--green);color:#fff}
.step-num{
  width:24px;height:24px;border-radius:50%;
  background:rgba(255,255,255,0.1);
  font-size:0.7rem;font-weight:600;font-family:var(--mono);
  display:flex;align-items:center;justify-content:center;
  flex-shrink:0;margin-top:1px;transition:background 0.2s;
}
.step-item.active .step-num{background:var(--accent);color:#fff}
.step-info{}
.step-title{font-size:0.85rem;font-weight:500;color:#f0efe8}
.step-item:not(.active) .step-title{color:#9a9888}
.step-desc{font-size:0.7rem;color:#5a5a50;margin-top:0.1rem;font-family:var(--mono)}

.sidebar-footer{
  padding-top:1.5rem;border-top:1px solid rgba(255,255,255,0.08);
  font-size:0.7rem;color:#4a4a42;font-family:var(--mono);
}

/* main */
.main{flex:1;padding:2.5rem 3rem;max-width:820px}

.page{display:none}
.page.active{display:block}

.page-header{margin-bottom:2rem}
.page-label{
  font-family:var(--mono);font-size:0.65rem;letter-spacing:0.15em;
  text-transform:uppercase;color:var(--accent);margin-bottom:0.5rem;
}
.page-title{font-size:1.8rem;font-weight:600;letter-spacing:-0.03em;line-height:1.2}
.page-sub{color:var(--muted);font-size:0.9rem;margin-top:0.4rem}

/* cards */
.card{
  background:var(--surface);border:1px solid var(--border);
  border-radius:10px;padding:1.5rem;margin-bottom:1rem;
}
.card-label{
  font-family:var(--mono);font-size:0.65rem;letter-spacing:0.12em;
  text-transform:uppercase;color:var(--muted);margin-bottom:1rem;
}

/* upload zone */
.upload-zone{
  border:2px dashed var(--border2);border-radius:8px;
  padding:3rem 2rem;text-align:center;cursor:pointer;
  transition:all 0.2s;position:relative;background:var(--surface2);
}
.upload-zone:hover,.upload-zone.drag{
  border-color:var(--accent);background:var(--accent-bg);
}
.upload-zone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%}
.upload-icon{font-size:2rem;margin-bottom:0.8rem}
.upload-zone h3{font-size:1rem;font-weight:600}
.upload-zone p{color:var(--muted);font-size:0.82rem;margin-top:0.3rem;font-family:var(--mono)}
.file-badge{
  display:inline-flex;align-items:center;gap:0.4rem;
  background:var(--green-bg);color:var(--green);
  border:1px solid #b8e8d0;border-radius:20px;
  font-size:0.78rem;padding:0.3rem 0.8rem;margin-top:0.8rem;
  font-family:var(--mono);
}

/* table */
.tbl-wrap{overflow-x:auto;border-radius:8px;border:1px solid var(--border)}
table{width:100%;border-collapse:collapse;font-size:0.8rem;font-family:var(--mono)}
th{
  background:var(--surface2);color:var(--muted);
  padding:0.55rem 0.9rem;text-align:left;font-size:0.65rem;
  letter-spacing:0.08em;text-transform:uppercase;border-bottom:1px solid var(--border);
}
td{padding:0.5rem 0.9rem;border-bottom:1px solid var(--border);color:var(--text)}
tr:last-child td{border-bottom:none}
tr:hover td{background:var(--surface2)}

/* col map */
.col-grid{display:grid;grid-template-columns:1fr 1fr;gap:0.8rem}
.col-item{display:flex;flex-direction:column;gap:0.3rem}
.col-item label{font-size:0.75rem;color:var(--muted);font-family:var(--mono)}
select,input[type=text]{
  background:var(--surface2);border:1px solid var(--border2);
  color:var(--text);padding:0.5rem 0.7rem;border-radius:6px;
  font-family:var(--mono);font-size:0.82rem;outline:none;
  transition:border-color 0.15s;width:100%;
}
select:focus,input[type=text]:focus{border-color:var(--accent)}

/* rubros selector */
.rubros-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:0.8rem}
.rubros-search{
  flex:1;max-width:260px;
  background:var(--surface2);border:1px solid var(--border2);
  border-radius:6px;padding:0.45rem 0.7rem;
  font-family:var(--mono);font-size:0.82rem;outline:none;color:var(--text);
}
.rubros-search:focus{border-color:var(--accent)}
.rubros-actions{display:flex;gap:0.5rem}
.btn-sm{
  font-size:0.75rem;padding:0.35rem 0.8rem;border-radius:5px;
  border:1px solid var(--border2);background:var(--surface2);
  color:var(--muted);cursor:pointer;font-family:var(--mono);
  transition:all 0.15s;
}
.btn-sm:hover{background:var(--border);color:var(--text)}

.rubros-grid{
  display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));
  gap:0.4rem;max-height:320px;overflow-y:auto;padding-right:4px;
}
.rubro-chip{
  display:flex;align-items:center;gap:0.5rem;
  padding:0.45rem 0.7rem;border-radius:6px;
  border:1px solid var(--border);background:var(--surface2);
  cursor:pointer;transition:all 0.15s;user-select:none;
  font-size:0.78rem;
}
.rubro-chip:hover{border-color:var(--border2);background:var(--border)}
.rubro-chip.selected{
  border-color:var(--accent);background:var(--accent-bg);color:var(--accent);font-weight:500;
}
.rubro-chip .dot{
  width:8px;height:8px;border-radius:50%;
  border:1.5px solid var(--border2);flex-shrink:0;transition:all 0.15s;
}
.rubro-chip.selected .dot{background:var(--accent);border-color:var(--accent)}
.rubro-count{
  font-family:var(--mono);font-size:0.65rem;
  color:var(--muted);margin-left:auto;
}
.selected-summary{
  font-family:var(--mono);font-size:0.75rem;color:var(--muted);
  margin-top:0.6rem;
}
.selected-summary span{color:var(--accent);font-weight:500}

/* modo */
.mode-grid{display:grid;grid-template-columns:1fr 1fr;gap:0.8rem}
.mode-card{
  border:2px solid var(--border);border-radius:8px;padding:1.2rem;
  cursor:pointer;transition:all 0.15s;
}
.mode-card:hover{border-color:var(--border2)}
.mode-card.selected{border-color:var(--accent);background:var(--accent-bg)}
.mode-icon{font-size:1.5rem;margin-bottom:0.5rem}
.mode-title{font-weight:600;font-size:0.9rem}
.mode-desc{font-size:0.78rem;color:var(--muted);margin-top:0.3rem;line-height:1.4}

/* buttons */
.btn{
  display:inline-flex;align-items:center;justify-content:center;gap:0.5rem;
  padding:0.75rem 1.5rem;border-radius:8px;border:none;
  font-family:var(--font);font-weight:600;font-size:0.9rem;
  cursor:pointer;transition:all 0.2s;letter-spacing:-0.01em;
}
.btn-primary{background:var(--accent);color:#fff}
.btn-primary:hover:not(:disabled){background:#b8420a;transform:translateY(-1px)}
.btn-primary:disabled{opacity:0.4;cursor:not-allowed}
.btn-ghost{background:transparent;border:1px solid var(--border2);color:var(--muted)}
.btn-ghost:hover{background:var(--surface2);color:var(--text)}
.btn-green{background:var(--green);color:#fff}
.btn-green:hover{background:#155a37;transform:translateY(-1px)}
.btn-row{display:flex;gap:0.8rem;margin-top:1.5rem}

/* progress */
.progress-wrap{margin:1.2rem 0}
.progress-track{
  background:var(--surface2);border-radius:100px;height:8px;
  overflow:hidden;border:1px solid var(--border);
}
.progress-fill{
  height:100%;border-radius:100px;
  background:linear-gradient(90deg,var(--accent),#f07030);
  transition:width 0.4s ease;
}
.progress-meta{
  display:flex;justify-content:space-between;
  font-family:var(--mono);font-size:0.72rem;color:var(--muted);margin-top:0.4rem;
}
.empresa-pill{
  display:inline-flex;align-items:center;gap:0.4rem;
  background:var(--blue-bg);color:var(--blue);
  border:1px solid #c0d8f0;border-radius:20px;
  font-size:0.75rem;padding:0.3rem 0.8rem;
  font-family:var(--mono);max-width:100%;overflow:hidden;
  text-overflow:ellipsis;white-space:nowrap;margin-top:0.6rem;
}

/* stats */
.stats-row{display:grid;grid-template-columns:repeat(4,1fr);gap:0.6rem;margin-bottom:1rem}
.stat-box{
  background:var(--surface);border:1px solid var(--border);
  border-radius:8px;padding:1rem;text-align:center;
}
.stat-num{font-size:1.6rem;font-weight:700;font-family:var(--mono);color:var(--accent)}
.stat-label{font-size:0.65rem;color:var(--muted);text-transform:uppercase;letter-spacing:0.08em;margin-top:0.2rem}

/* logs */
.logbox{
  background:#1a1916;border-radius:8px;padding:1rem;
  height:200px;overflow-y:auto;font-family:var(--mono);font-size:0.72rem;
}
.log-line{color:#5a5a50;margin-bottom:0.15rem;line-height:1.5}
.log-line.ok{color:#4caf80}
.log-line.err{color:#e05050}
.log-ts{color:#3a3a32;margin-right:0.5rem}

/* done */
.done-box{
  background:var(--green-bg);border:1px solid #b8e8d0;
  border-radius:10px;padding:1.5rem;text-align:center;margin-bottom:1.2rem;
}
.done-box h2{font-size:1.3rem;color:var(--green)}
.done-box p{font-size:0.85rem;color:var(--muted);font-family:var(--mono);margin-top:0.3rem}

/* scrollbar */
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:10px}

@media(max-width:700px){
  .shell{flex-direction:column}
  .sidebar{width:100%;height:auto;flex-direction:row;flex-wrap:wrap;padding:1rem;gap:0.5rem}
  .steps{display:flex;flex-wrap:wrap;gap:0.3rem}
  .logo-sub,.sidebar-footer{display:none}
  .main{padding:1.2rem}
  .stats-row{grid-template-columns:repeat(2,1fr)}
  .col-grid,.mode-grid{grid-template-columns:1fr}
}
</style>
</head>
<body>
<div class="shell">

<!-- SIDEBAR -->
<aside class="sidebar">
  <div class="logo">Lead<span>Enricher</span></div>
  <div class="logo-sub">Perú · v1.0</div>
  <nav class="steps" id="stepNav">
    <div class="step-item active" data-step="1" onclick="goStep(1)">
      <div class="step-num">1</div>
      <div class="step-info">
        <div class="step-title">Fuente de datos</div>
        <div class="step-desc">CSV o rubros CCL</div>
      </div>
    </div>
    <div class="step-item" data-step="2" onclick="goStep(2)">
      <div class="step-num">2</div>
      <div class="step-info">
        <div class="step-title">Seleccionar leads</div>
        <div class="step-desc">Revisar y filtrar</div>
      </div>
    </div>
    <div class="step-item" data-step="3" onclick="goStep(3)">
      <div class="step-num">3</div>
      <div class="step-info">
        <div class="step-title">Configurar</div>
        <div class="step-desc">Mapear columnas</div>
      </div>
    </div>
    <div class="step-item" data-step="4" onclick="goStep(4)">
      <div class="step-num">4</div>
      <div class="step-info">
        <div class="step-title">Procesar</div>
        <div class="step-desc">Enriquecer datos</div>
      </div>
    </div>
  </nav>
  <div class="sidebar-footer">
    SUNAT · PáginasAmarillas<br>DuckDuckGo · CCL
  </div>
</aside>

<!-- MAIN -->
<main class="main">

  <!-- PASO 1: FUENTE -->
  <div class="page active" id="page1">
    <div class="page-header">
      <div class="page-label">Paso 01</div>
      <div class="page-title">¿De dónde vienen<br>tus empresas?</div>
      <div class="page-sub">Sube tu propio Excel/CSV o elige rubros del directorio CCL</div>
    </div>

    <div class="mode-grid">
      <div class="mode-card selected" id="modeCSV" onclick="selectMode('csv')">
        <div class="mode-icon">📂</div>
        <div class="mode-title">Subir mi CSV / Excel</div>
        <div class="mode-desc">Ya tengo una lista de empresas. Quiero enriquecerla con RUC, gerente, teléfono y dirección.</div>
      </div>
      <div class="mode-card" id="modeRubros" onclick="selectMode('rubros')">
        <div class="mode-icon">🏷️</div>
        <div class="mode-title">Elegir rubros CCL</div>
        <div class="mode-desc">Quiero scrapear el directorio de la Cámara de Comercio de Lima por rubro y enriquecer esas empresas.</div>
      </div>
    </div>

    <!-- CSV upload -->
    <div id="csvBlock" style="margin-top:1rem">
      <div class="card">
        <div class="card-label">Sube tu archivo</div>
        <div class="upload-zone" id="dropZone">
          <input type="file" id="fileInput" accept=".csv,.xlsx,.xls">
          <div class="upload-icon">📄</div>
          <h3>Arrastra tu CSV o Excel aquí</h3>
          <p>Columnas mínimas: <strong>empresa</strong> o <strong>nombre</strong></p>
          <div id="fileName"></div>
        </div>
      </div>
      <div class="btn-row" style="margin-top:0">
        <button class="btn btn-primary" id="btnCsvNext" disabled onclick="goStep(2)">Continuar →</button>
      </div>
    </div>

    <!-- Rubros picker -->
    <div id="rubrosBlock" style="display:none;margin-top:1rem">
      <div class="card">
        <div class="card-label">Rubros disponibles en CCL</div>
        <div class="rubros-header">
          <input class="rubros-search" id="rubrosSearch" placeholder="Buscar rubro..." oninput="filtrarRubros(this.value)">
          <div class="rubros-actions">
            <button class="btn-sm" onclick="selectAllRubros()">Todos</button>
            <button class="btn-sm" onclick="clearRubros()">Ninguno</button>
          </div>
        </div>
        <div class="rubros-grid" id="rubrosGrid">
          <div style="color:var(--muted);font-family:var(--mono);font-size:0.8rem;padding:1rem">
            Cargando rubros...
          </div>
        </div>
        <div class="selected-summary" id="rubrosSummary">0 rubros seleccionados</div>
      </div>
      <div class="btn-row" style="margin-top:0">
        <button class="btn btn-primary" id="btnRubrosNext" disabled onclick="iniciarScrapeRubros()">Scrapear y continuar →</button>
      </div>
    </div>
  </div>

  <!-- PASO 2: SELECCIONAR LEADS -->
  <div class="page" id="page2">
    <div class="page-header">
      <div class="page-label">Paso 02</div>
      <div class="page-title">Revisar leads</div>
      <div class="page-sub">Elige cuáles empresas procesar. Puedes deseleccionar las que no quieras.</div>
    </div>

    <div class="card">
      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:0.8rem">
        <div class="card-label" style="margin:0">Vista previa de empresas</div>
        <div style="display:flex;gap:0.5rem;align-items:center">
          <input id="leadSearch" style="width:200px;padding:0.4rem 0.6rem" type="text" placeholder="Filtrar..." oninput="filtrarLeads(this.value)">
          <button class="btn-sm" onclick="selectAllLeads()">Todas</button>
          <button class="btn-sm" onclick="clearLeads()">Ninguna</button>
        </div>
      </div>
      <div class="tbl-wrap" style="max-height:380px;overflow-y:auto">
        <table id="leadsTable">
          <thead><tr id="leadsHeader"></tr></thead>
          <tbody id="leadsBody"></tbody>
        </table>
      </div>
      <div style="font-family:var(--mono);font-size:0.75rem;color:var(--muted);margin-top:0.6rem">
        <span id="leadsCount">0</span> empresas seleccionadas
      </div>
    </div>

    <div class="btn-row">
      <button class="btn btn-ghost" onclick="goStep(1)">← Volver</button>
      <button class="btn btn-primary" id="btnLeadsNext" onclick="goStep(3)">Configurar enriquecimiento →</button>
    </div>
  </div>

  <!-- PASO 3: CONFIGURAR -->
  <div class="page" id="page3">
    <div class="page-header">
      <div class="page-label">Paso 03</div>
      <div class="page-title">Configurar</div>
      <div class="page-sub">Dile al sistema qué columna es cada cosa</div>
    </div>

    <div class="card">
      <div class="card-label">Mapeo de columnas</div>
      <div class="col-grid" id="colMapGrid"></div>
    </div>

    <div class="card">
      <div class="card-label">¿Qué datos buscar?</div>
      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:0.6rem" id="datosChk">
        <label style="display:flex;align-items:center;gap:0.5rem;font-size:0.85rem;cursor:pointer">
          <input type="checkbox" id="chkRuc" checked> RUC
        </label>
        <label style="display:flex;align-items:center;gap:0.5rem;font-size:0.85rem;cursor:pointer">
          <input type="checkbox" id="chkRep" checked> Representante legal
        </label>
        <label style="display:flex;align-items:center;gap:0.5rem;font-size:0.85rem;cursor:pointer">
          <input type="checkbox" id="chkDir" checked> Dirección fiscal
        </label>
        <label style="display:flex;align-items:center;gap:0.5rem;font-size:0.85rem;cursor:pointer">
          <input type="checkbox" id="chkTel" checked> Teléfono
        </label>
        <label style="display:flex;align-items:center;gap:0.5rem;font-size:0.85rem;cursor:pointer">
          <input type="checkbox" id="chkEmail" checked> Email
        </label>
      </div>
    </div>

    <div class="btn-row">
      <button class="btn btn-ghost" onclick="goStep(2)">← Volver</button>
      <button class="btn btn-primary" onclick="goStep(4);iniciarEnriquecimiento()">⚡ Iniciar enriquecimiento</button>
    </div>
  </div>

  <!-- PASO 4: PROCESAR -->
  <div class="page" id="page4">
    <div class="page-header">
      <div class="page-label">Paso 04</div>
      <div class="page-title" id="p4Title">Procesando...</div>
      <div class="page-sub" id="p4Sub">Consultando SUNAT, Páginas Amarillas y más fuentes</div>
    </div>

    <div class="stats-row">
      <div class="stat-box"><div class="stat-num" id="st-total">0</div><div class="stat-label">Procesadas</div></div>
      <div class="stat-box"><div class="stat-num" id="st-rep">0</div><div class="stat-label">Con Gerente</div></div>
      <div class="stat-box"><div class="stat-num" id="st-tel">0</div><div class="stat-label">Con Teléfono</div></div>
      <div class="stat-box"><div class="stat-num" id="st-email">0</div><div class="stat-label">Con Email</div></div>
    </div>

    <div class="card" id="progressCard">
      <div class="progress-wrap">
        <div class="progress-track">
          <div class="progress-fill" id="progFill" style="width:0%"></div>
        </div>
        <div class="progress-meta">
          <span id="progText">0 / 0 empresas</span>
          <span id="progPct">0%</span>
        </div>
      </div>
      <div id="empPill" class="empresa-pill" style="display:none">⚡ <span id="empName"></span></div>
    </div>

    <div class="card">
      <div class="card-label">Log en vivo</div>
      <div class="logbox" id="logBox"></div>
    </div>

    <!-- Done -->
    <div id="doneBlock" style="display:none">
      <div class="done-box">
        <h2>✅ ¡Listo!</h2>
        <p id="doneMsg"></p>
      </div>
      <button class="btn btn-green" style="width:100%;font-size:1rem;padding:1rem" onclick="descargar()">
        ⬇️ Descargar CSV enriquecido
      </button>
    </div>
  </div>

</main>
</div>

<script>
// ── STATE ─────────────────────────────────────────────────────────────────────
let mode = 'csv';
let csvRows = [], csvHeaders = [];
let allLeads = [];        // todas las empresas cargadas
let selectedLeads = [];   // las que el usuario seleccionó (checkbox)
let rubrosMap = {};       // label -> id  (de CCL)
let rubrosSeleccionados = new Set();
let colMap = { empresa: '', rubro: '', ruc: '' };
let logCount = 0;
let pollTimer = null;
let currentStep = 1;

// ── NAVEGACIÓN ────────────────────────────────────────────────────────────────
function goStep(n) {
  if (n > currentStep + 1) return;
  currentStep = n;
  if (n===3) setTimeout(renderColMap, 50);
  document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
  document.getElementById('page' + n).classList.add('active');
  document.querySelectorAll('.step-item').forEach(s => {
    const sn = parseInt(s.dataset.step);
    s.classList.remove('active','done');
    if (sn === n) s.classList.add('active');
    if (sn < n)  s.classList.add('done');
    const dot = s.querySelector('.step-num');
    if (sn < n) dot.textContent = '✓';
    else dot.textContent = sn;
  });
}

// ── MODO ─────────────────────────────────────────────────────────────────────
function selectMode(m) {
  mode = m;
  document.getElementById('modeCSV').classList.toggle('selected', m==='csv');
  document.getElementById('modeRubros').classList.toggle('selected', m==='rubros');
  document.getElementById('csvBlock').style.display = m==='csv' ? 'block' : 'none';
  document.getElementById('rubrosBlock').style.display = m==='rubros' ? 'block' : 'none';
  if (m==='rubros' && Object.keys(rubrosMap).length === 0) cargarRubros();
}

// ── CSV ───────────────────────────────────────────────────────────────────────
document.getElementById('fileInput').addEventListener('change', e => {
  if (e.target.files[0]) handleFile(e.target.files[0]);
});
const dz = document.getElementById('dropZone');
dz.addEventListener('dragover', e => { e.preventDefault(); dz.classList.add('drag'); });
dz.addEventListener('dragleave', () => dz.classList.remove('drag'));
dz.addEventListener('drop', e => {
  e.preventDefault(); dz.classList.remove('drag');
  if (e.dataTransfer.files[0]) handleFile(e.dataTransfer.files[0]);
});

function handleFile(file) {
  console.log('[handleFile] Archivo:', file.name, file.size, file.type);
  document.getElementById('fileName').innerHTML =
    `<div class="file-badge">⏳ Subiendo ${file.name}...</div>`;

  // Subir al servidor para que Python lo parsee (soporta CSV y XLSX)
  const fd = new FormData();
  fd.append('file', file);
  fetch('/upload', { method: 'POST', body: fd })
    .then(r => r.json())
    .then(d => {
      if (d.error) { alert('Error: ' + d.error); return; }
      csvHeaders = d.headers;
      csvRows    = d.rows;
      allLeads   = csvRows.map(r => ({...r, _sel: true}));
      guessColMap();
      document.getElementById('fileName').innerHTML =
        `<div class="file-badge">✅ ${file.name} · ${d.rows.length} filas</div>`;
      document.getElementById('btnCsvNext').disabled = false;
      renderLeadsTable();
    })
    .catch(e => {
      document.getElementById('fileName').innerHTML =
        `<div class="file-badge" style="color:red">❌ Error subiendo archivo</div>`;
    });
}

function guessColMap() {
  const g = (kws) => csvHeaders.find(h => kws.some(k => h.toLowerCase().includes(k))) || '';
  colMap.empresa = g(['empresa','nombre','razon','company','name']);
  colMap.rubro   = g(['rubro','categoria','sector','industria','giro']);
  colMap.ruc     = g(['ruc','documento','doc']);
}

// ── LEADS TABLE ───────────────────────────────────────────────────────────────
function renderLeadsTable(filter='') {
  const filt = filter.toLowerCase();
  const visible = allLeads.filter(r => {
    if (!filt) return true;
    return Object.values(r).some(v => String(v).toLowerCase().includes(filt));
  });

  const hdr = document.getElementById('leadsHeader');
  const body = document.getElementById('leadsBody');

  const cols = csvHeaders.length ? csvHeaders : ['empresa','rubro'];
  hdr.innerHTML = `<th style="width:32px"><input type="checkbox" id="chkAll" onchange="toggleAll(this.checked)" checked></th>` +
    cols.map(c=>`<th>${c}</th>`).join('');

  body.innerHTML = visible.map((row, i) => {
    const idx = allLeads.indexOf(row);
    return `<tr>
      <td><input type="checkbox" data-idx="${idx}" ${row._sel?'checked':''} onchange="toggleLead(${idx},this.checked)"></td>
      ${cols.map(c=>`<td>${row[c]||''}</td>`).join('')}
    </tr>`;
  }).join('');

  updateLeadsCount();
}

function filtrarLeads(q) { renderLeadsTable(q); }
function toggleLead(idx, checked) { allLeads[idx]._sel = checked; updateLeadsCount(); }
function toggleAll(checked) { allLeads.forEach(r => r._sel = checked); renderLeadsTable(); }
function selectAllLeads() { allLeads.forEach(r => r._sel = true); renderLeadsTable(); }
function clearLeads() { allLeads.forEach(r => r._sel = false); renderLeadsTable(); }
function updateLeadsCount() {
  const n = allLeads.filter(r=>r._sel).length;
  document.getElementById('leadsCount').textContent = n;
  document.getElementById('btnLeadsNext').disabled = n === 0;
}

// ── COL MAP (paso 3) ──────────────────────────────────────────────────────────
function renderColMap() {
  const fields = [
    {key:'empresa', label:'🏢 Columna Empresa *'},
    {key:'rubro',   label:'🏷️ Columna Rubro'},
    {key:'ruc',     label:'🔢 Columna RUC (opcional)'},
  ];
  const hdrs = csvHeaders.length ? csvHeaders : Object.keys(allLeads[0]||{}).filter(k=>k!=='_sel');
  document.getElementById('colMapGrid').innerHTML = fields.map(f=>`
    <div class="col-item">
      <label>${f.label}</label>
      <select onchange="colMap['${f.key}']=this.value">
        <option value="">— ninguna —</option>
        ${hdrs.map(h=>`<option value="${h}" ${colMap[f.key]===h?'selected':''}>${h}</option>`).join('')}
      </select>
    </div>
  `).join('');
}

// renderColMap se llama desde dentro de goStep directamente
// (no override — evita recursión infinita)

// ── RUBROS CCL ────────────────────────────────────────────────────────────────
async function cargarRubros() {
  const grid = document.getElementById('rubrosGrid');
  grid.innerHTML = '<div style="color:var(--muted);font-family:var(--mono);font-size:0.8rem;padding:1rem">Cargando rubros de CCL...</div>';
  try {
    const r = await fetch('/rubros');
    rubrosMap = await r.json();
    renderRubrosGrid(Object.keys(rubrosMap));
  } catch(e) {
    grid.innerHTML = '<div style="color:#e05050;font-family:var(--mono);font-size:0.8rem;padding:1rem">Error cargando rubros</div>';
  }
}

function renderRubrosGrid(rubros) {
  const grid = document.getElementById('rubrosGrid');
  grid.innerHTML = rubros.map(r => `
    <div class="rubro-chip ${rubrosSeleccionados.has(r)?'selected':''}" onclick="toggleRubro('${r.replace(/'/g,"\\'")}')">
      <div class="dot"></div>
      <span style="flex:1;line-height:1.3">${r}</span>
    </div>
  `).join('');
  updateRubrosSummary();
}

function toggleRubro(r) {
  if (rubrosSeleccionados.has(r)) rubrosSeleccionados.delete(r);
  else rubrosSeleccionados.add(r);
  renderRubrosGrid(Object.keys(rubrosMap));
  document.getElementById('btnRubrosNext').disabled = rubrosSeleccionados.size === 0;
}

function filtrarRubros(q) {
  const filtered = Object.keys(rubrosMap).filter(r => r.toLowerCase().includes(q.toLowerCase()));
  renderRubrosGrid(filtered);
}

function selectAllRubros() {
  Object.keys(rubrosMap).forEach(r => rubrosSeleccionados.add(r));
  renderRubrosGrid(Object.keys(rubrosMap));
  document.getElementById('btnRubrosNext').disabled = false;
}

function clearRubros() {
  rubrosSeleccionados.clear();
  renderRubrosGrid(Object.keys(rubrosMap));
  document.getElementById('btnRubrosNext').disabled = true;
}

function updateRubrosSummary() {
  const el = document.getElementById('rubrosSummary');
  el.innerHTML = `<span>${rubrosSeleccionados.size}</span> rubros seleccionados`;
}

async function iniciarScrapeRubros() {
  const btn = document.getElementById('btnRubrosNext');
  btn.disabled = true;
  btn.textContent = 'Scrapeando CCL...';
  try {
    const r = await fetch('/scrape-rubros', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({rubros: [...rubrosSeleccionados]})
    });
    const d = await r.json();
    if (d.empresas) {
      allLeads = d.empresas.map(e => ({...e, _sel: true}));
      csvHeaders = ['empresa','rubro'];
      colMap = { empresa:'empresa', rubro:'rubro', ruc:'' };
      renderLeadsTable();
      goStep(2);
    }
  } catch(e) {
    btn.textContent = 'Error, reintentar';
    btn.disabled = false;
  }
}

// ── ENRIQUECIMIENTO ───────────────────────────────────────────────────────────
async function iniciarEnriquecimiento() {
  selectedLeads = allLeads.filter(r => r._sel).map(r => {
    const obj = {...r};
    delete obj._sel;
    return {
      empresa: obj[colMap.empresa] || obj.empresa || '',
      rubro:   obj[colMap.rubro]   || obj.rubro   || '',
      ruc:     obj[colMap.ruc]     || obj.ruc      || '',
    };
  }).filter(e => e.empresa);

  logCount = 0;
  document.getElementById('logBox').innerHTML = '';

  const r = await fetch('/enriquecer', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ empresas: selectedLeads })
  });
  const d = await r.json();
  if (d.success) startPolling();
  else addLog('Error: ' + d.error, true);
}

function startPolling() {
  pollTimer = setInterval(async () => {
    try {
      const r = await fetch('/status');
      const d = await r.json();
      updateUI(d);
      if (d.finished) { clearInterval(pollTimer); showDone(d); }
    } catch {}
  }, 1000);
}

function updateUI(d) {
  const pct = d.progress || 0;
  document.getElementById('progFill').style.width = pct + '%';
  document.getElementById('progPct').textContent  = pct + '%';
  document.getElementById('progText').textContent = `${d.index||0} / ${d.total||0} empresas`;
  document.getElementById('st-total').textContent = d.index  || 0;
  document.getElementById('st-rep').textContent   = d.con_rep || 0;
  document.getElementById('st-tel').textContent   = d.con_tel || 0;
  document.getElementById('st-email').textContent = d.con_email || 0;

  if (d.empresa_actual) {
    document.getElementById('empPill').style.display = 'inline-flex';
    document.getElementById('empName').textContent = d.empresa_actual;
  }

  const logs = d.logs || [];
  logs.slice(logCount).forEach(l => {
    addLog(l.msg, l.err, l.ok);
    logCount++;
  });
}

function addLog(msg, err=false, ok=false) {
  const box = document.getElementById('logBox');
  const d = document.createElement('div');
  d.className = 'log-line' + (ok?' ok':err?' err':'');
  const t = new Date().toLocaleTimeString('es-PE',{hour12:false});
  d.innerHTML = `<span class="log-ts">${t}</span>${msg}`;
  box.appendChild(d);
  box.scrollTop = box.scrollHeight;
}

function showDone(d) {
  document.getElementById('p4Title').textContent = '¡Completado!';
  document.getElementById('p4Sub').textContent   = 'Listo para descargar';
  document.getElementById('progressCard').style.display = 'none';
  document.getElementById('doneBlock').style.display = 'block';
  const secs = d.finished_at && d.started_at ? Math.round(d.finished_at - d.started_at) : 0;
  document.getElementById('doneMsg').textContent =
    `${d.total} empresas · ${d.con_rep} con gerente · ${d.con_tel} con teléfono · ${secs}s`;
  document.getElementById('st-total').textContent = d.total    || 0;
  document.getElementById('st-rep').textContent   = d.con_rep  || 0;
  document.getElementById('st-tel').textContent   = d.con_tel  || 0;
  document.getElementById('st-email').textContent = d.con_email|| 0;
}

function descargar() { window.location = '/download'; }
</script>
</body>
</html>
"""

# ─── BACKGROUND ──────────────────────────────────────────────────────────────

def _log(msg, ok=False, err=False):
    with _LOCK:
        STATE["logs"].append({"ts": time.time(), "msg": msg, "ok": ok, "err": err})
        STATE["logs"] = STATE["logs"][-120:]

def _progress_cb(evt):
    with _LOCK:
        STATE["index"]          = evt.get("index", 0)
        STATE["total"]          = evt.get("total", 0)
        STATE["progress"]       = evt.get("progress", 0)
        STATE["empresa_actual"] = evt.get("empresa", "")
        STATE["con_rep"]        = evt.get("con_rep", 0)
        STATE["con_tel"]        = evt.get("con_tel", 0)
        STATE["con_email"]      = evt.get("con_email", 0)
    status = evt.get("status", "ok")
    empresa = evt.get("empresa", "")
    rep = evt.get("con_rep", 0)
    tel = evt.get("con_tel", 0)
    _log(f"{empresa}", ok=(status=="ok"), err=(status=="error"))

def background_enrich(empresas):
    global STATE
    try:
        results = enriquecer_lote(empresas, progress_cb=_progress_cb)
        with _LOCK:
            STATE["results"]     = results
            STATE["finished"]    = True
            STATE["finished_at"] = time.time()
            STATE["progress"]    = 100
            STATE["con_rep"]     = sum(1 for r in results if r.get("representante"))
            STATE["con_tel"]     = sum(1 for r in results if r.get("telefono"))
            STATE["con_email"]   = sum(1 for r in results if r.get("email"))
        _log(f"✅ Completado: {len(results)} empresas", ok=True)
    except Exception as e:
        with _LOCK:
            STATE["error"] = str(e)
            STATE["finished"] = True
        _log(f"❌ Error: {e}", err=True)
    finally:
        with _LOCK:
            STATE["running"] = False

# ─── ROUTES ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return HTML

@app.route("/rubros")
def rubros():
    try:
        from enricher import get_rubros_ccl
        import requests as req
        s = req.Session()
        s.headers.update({"User-Agent": "Mozilla/5.0"})
        rmap = get_rubros_ccl(s)
        return jsonify(rmap)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/scrape-rubros", methods=["POST"])
def scrape_rubros():
    from enricher import scrape_rubros_ccl
    data = request.get_json(silent=True) or {}
    rubros_lista = data.get("rubros", [])
    try:
        empresas = scrape_rubros_ccl(rubros_lista)
        return jsonify({"empresas": empresas})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/enriquecer", methods=["POST"])
def enriquecer():
    global STATE
    with _LOCK:
        if STATE.get("running"):
            return jsonify({"success": False, "error": "Ya hay un proceso corriendo"}), 409
    data = request.get_json(silent=True) or {}
    empresas = data.get("empresas", [])
    if not empresas:
        return jsonify({"success": False, "error": "Sin empresas"}), 400
    with _LOCK:
        STATE.update({
            "running": True, "finished": False, "error": None,
            "progress": 0, "index": 0, "total": len(empresas),
            "empresa_actual": "", "con_rep": 0, "con_tel": 0, "con_email": 0,
            "logs": [], "results": [], "started_at": time.time(), "finished_at": None,
        })
    _log(f"Iniciando: {len(empresas)} empresas...")
    threading.Thread(target=background_enrich, args=(empresas,), daemon=True).start()
    return jsonify({"success": True})

@app.route("/status")
def status():
    with _LOCK:
        return jsonify({k: v for k, v in STATE.items() if k != "results"})

@app.route("/download")
def download():
    with _LOCK:
        results = list(STATE.get("results", []))
    if not results:
        return jsonify({"error": "Sin resultados"}), 404
    from enricher import CAMPOS_SALIDA
    fields = CAMPOS_SALIDA
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=fields, extrasaction="ignore")
    w.writeheader()
    w.writerows(results)
    out.seek(0)
    return send_file(
        io.BytesIO(out.getvalue().encode("utf-8")),
        mimetype="text/csv", as_attachment=True,
        download_name="empresas_enriquecidas.csv",
    )


@app.route("/upload", methods=["POST"])
def upload():
    """Parsea CSV o XLSX en el servidor y devuelve JSON con headers+rows"""
    print(f"[UPLOAD] Request recibido. Files: {list(request.files.keys())}")
    f = request.files.get("file")
    if not f:
        print("[UPLOAD] ERROR: Sin archivo en request")
        return jsonify({"error": "Sin archivo"}), 400
    print(f"[UPLOAD] Archivo recibido: {f.filename}, content_type: {f.content_type}")

    filename = f.filename.lower()
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            import openpyxl, io as _io2
            file_bytes = f.read()
            print(f"[UPLOAD] Bytes leidos: {len(file_bytes)}")
            wb = openpyxl.load_workbook(_io2.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            print(f"[UPLOAD] Sheet: {ws.title}, max_row={ws.max_row}")
            all_rows = list(ws.iter_rows(values_only=True))
            print(f"[UPLOAD] Filas totales: {len(all_rows)}")
            if not all_rows:
                return jsonify({"error": "Archivo vacio"}), 400
            headers = [str(h).strip() if h is not None else f"col{i}" for i,h in enumerate(all_rows[0])]
            print(f"[UPLOAD] Headers: {headers[:5]}")
            rows = []
            for row in all_rows[1:]:
                obj = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        obj[headers[i]] = str(val).strip() if val is not None else ""
                if any(v for v in obj.values()):
                    rows.append(obj)
            print(f"[UPLOAD] Filas con datos: {len(rows)}")
        else:
            # CSV
            import io as _io
            text = f.read().decode("utf-8", errors="replace")
            lines = text.strip().split("\n")
            headers = [h.replace('"','').strip() for h in lines[0].split(",")]
            rows = []
            for line in lines[1:]:
                vals, cur, inq = [], "", False
                for c in line:
                    if c == '"': inq = not inq; continue
                    if c == "," and not inq: vals.append(cur.strip()); cur = ""; continue
                    cur += c
                vals.append(cur.strip())
                obj = {headers[i]: vals[i] if i < len(vals) else "" for i in range(len(headers))}
                if any(v for v in obj.values()):
                    rows.append(obj)

        return jsonify({"headers": headers, "rows": rows, "total": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True, port=5000)
